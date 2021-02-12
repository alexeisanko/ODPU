import openpyxl
import pandas as pd
import os
import sys
import gc
import time
import csv
import colorama
colorama.init()


def main_file():
    """Функция открывает файл для заполнения, создает и подготоавливает
     из него массив из списка городов для обработки данных"""

    # Запрашиваем путь к файлу baza.xlsx
    print('Начата подготовка файла для внесения информации')

    # Создаем массив из файла на основе листа 'Правильный порядок'
    base = openpyxl.load_workbook(os.path.abspath('baza.xlsx'))
    roster_town = pd.DataFrame(base['правильный порядок'].values)

    # Переносим первую строчку массива в шапку(название)
    for i in range(len(roster_town.columns)):
        roster_town = roster_town.rename(columns={roster_town.columns[i]: roster_town.iloc[0, i]})
    roster_town = roster_town[1:]

    # Преобразуем строки к единому стилю (без ъ, ё и в нижнем регистре)
    for i in range(len(roster_town.index)):
        if roster_town.iloc[i, 0] is None:
            continue

        # Для правильного учета обьемов столицы
        if roster_town.iloc[i, 1] == 'ГО "Сыктывкар"':
            roster_town.iloc[i, 1] = 'сыктывкар'

        roster_town.iloc[i, 0] = change_line(roster_town.iloc[i, 0])
        roster_town.iloc[i, 1] = change_line(roster_town.iloc[i, 1])
        roster_town.iloc[i, 2] = change_line(roster_town.iloc[i, 2])

    # Подготовка перечня городов для заполнения
    roster_town = change_base(roster_town)
    roster_town = find_district_center(roster_town)
    return roster_town


def change_line(line: str, user=False):
    """Функция заменяет в строке буквы 'ъ, ё' на 'ь, е
    а также приводит строку к нижнему регистру"""

    change = str.maketrans('ъё', 'ье')
    new_line = line.translate(change)
    new_line = new_line.lower()
    if user:
        pass
    return new_line


def change_base(base):
    """Функция подготавливает базу для заполнения. Учитывает все
    сокрашенные префиксы у населенных пунктов, создает множество для
    учета всех уникальных номеров счетчиков"""

    # Учитываем все префиксы
    prefix = {'г': ['г', 'г.', 'город', 'гор', 'гор.'],
              'п': ['п', 'п.', 'пос', 'пос.', 'поселок'],
              'д': ['д', 'д.', 'дер', 'дер.', 'деревня'],
              'с': ['с', 'с.', 'село'],
              'пгт': ['пгт']}

    # Перебираем все населенные пункты и вносим все префиксы в базу
    for i in range(len(base.index)):
        if base.iloc[i, 0] is None:
            continue
        locality = base.iloc[i, 2].split()
        new_locality = []
        for pref in prefix[locality[0]]:
            new_locality.append(pref + ' ' + ' '.join(locality[1:]))
        base.iloc[i, 7] = new_locality

        # Убираем изначальный префикс нас. пунктов
        base.iloc[i, 2] = ' '.join(locality[1:])

        # int для расчета обьемов потребления
        base.iloc[i, 4] = 0
        base.iloc[i, 6] = 0

        # Множества для уникальных номеров счетчиков
        base.iloc[i, 3] = set()
        base.iloc[i, 5] = set()
    return base


def find_district_center(base):
    """Функция находит районный центр что бы не учитывать у него
    прилежащие районые пункты"""

    # Перебор населенных пунктов
    for i in range(len(base.index)):
        if base.iloc[i, 0] is None:
            continue

        # Проверка принадлежности населенного пункта к центру
        if base.iloc[i, 1] == base.iloc[i, 2]:
            base.iloc[i, 2] = []
            for m in range(i + 1, len(base.index)):
                if (base.iloc[i, 0] is None) or \
                        (base.iloc[m, 1] == base.iloc[m, 2]) or \
                        (base.iloc[m - 1, 1] != base.iloc[m, 1]):
                    break
                elif base.iloc[m, 2] in base.iloc[m, 0]:
                    pass
                else:
                    base.iloc[i, 2].append(base.iloc[m, 2])
    return base


def find_files_with_information():
    """Ищет в указанной папке все файлы с необходимой информацией"""
    files = []
    catalog = input(colorama.Style.BRIGHT + colorama.Fore.GREEN + 'Укажите путь к каталогу с данными: ')
    print(colorama.Style.RESET_ALL)
    # Перебираем все файлы в каталоге
    for name_file in os.listdir(catalog):
        if 'xlsx' in name_file and \
                ('Ведомость' in name_file or 'ведомость' in name_file):
            name_file = catalog + '/' + name_file
            files.append(name_file)
    return files


def import_information(base, file_with_info, unreported):
    """Открывает файл с информацией после чего преобразовывает,
    считывает и заносит данные в файл baza.xlsx"""

    # Открыть нужный Excel файл с информацией и считать каждый лист
    print("Загрузка файла " + file_with_info.split('/')[-1])
    work_book = openpyxl.load_workbook(file_with_info, read_only=True)
    for sheet in work_book:
        gc.collect()
        print('Считывается информация с загруженного файла, Лист: ' + sheet.title)
        # Если у листа мало столбцев, то есть пропускаем
        if sheet.max_column < 40:
            continue
        if sheet.max_column > 80:
            data = excel_to_csv(sheet)
        else:
            data = pd.DataFrame(sheet.values)
        data = optimization_data(data)

        # Если неправильный лист прошел отбор
        if type(data) == str:
            print(colorama.Fore.RED + 'Error. Wrong list')
            print('Ошибка считывания информации.')
            print('Возможно программа ощибочно пытается считать файл {}, Лист {}.'.format(file_with_info.split('/')[-1],
                                                                                          sheet.title))
            act = input(colorama.Style.BRIGHT +
                        'Просмотрите это лист. если программа ощибочно считывает его, введите "да" для его пропуска\n'
                        'либо введите "нет" для завершения программы\n')
            while act != 'да' and act != 'Да' and act != 'ДА' and act != 'нет' and act != 'Нет' and act != 'НЕТ':
                print('Некорректно введено необходимое действие')
                act = input(
                    'Просмотрите это лист. если программа ощибочно считывает его, введите "да" для его пропуска\n'
                    'либо введите "нет" для завершения программы\n'
                    '')
            else:
                if act == 'да' or act == 'да' or act == 'ДА':
                    continue
                elif act == 'нет' or act == 'Нет' or act == 'НЕТ':
                    print('Программа завершит свою работу через 5 секунд')
                    time.sleep(5)
                    sys.exit()
            print(colorama.Style.RESET_ALL)
        base, unreported = enter_information(base, data, unreported)
    return base, unreported


def excel_to_csv(sheet_in_excel):
    """Для возможности считывания больших файлов слабым компьютером, осущесвляет перевод формата xlsx в csv"""
    with open('myfile.csv', 'w') as out:
        writer = csv.writer(out)
        for row in sheet_in_excel:
            values = (cell.value for cell in row[:71])
            writer.writerow(values)
    sheet_in_csv = pd.read_csv('myfile.csv', dtype=object)
    return sheet_in_csv


def optimization_data(data):
    """Функция убирает лишние (пустые строки) и именнует столбцы.
     После чего фильтрует данные и оставляет только нужные столбцы (РЭС, Адрес, Объемы, № ПУ.
     Также при необходимости фильтрует тип потребителя (население и прриравненное к населению)"""

    for i, j in brute_force(len(data.index), len(data.columns)):
        if i > 15:
            return "Error. Wrong list"
        elif type(data.iloc[i, j]) is not str:
            continue
        elif 'Адрес' in data.iloc[i, j]:
            data = data[i:]
            break
    for i in range(len(data.columns)):
        data = data.rename(columns={data.columns[i]: data.iloc[0, i]})
    data = data[2:]
    necessary_columns = {'physical_users': ['РЭС', 'Адрес', 'Номер счетчика',
                                            'Объем \nпереданных расходов ГП за текущий период'],
                         'legal_users': ['РЭС', 'Адрес объекта', '№ ПУ', 'Общ расход']}
    users = ['Приравненные к городскому населению кроме эл.плит',
             'Приравненные к городскому населению (с эл.плитами)',
             'Население и приравненные к нему (городское без эл.плит)',
             'Население сельское',
             'Приравненные к сельскому населению',
             'Население городское (с эл.плитами)'
             ]
    if len(set(necessary_columns['physical_users']) & set(data)) == 4:
        data = pd.DataFrame([pd.Series(data[necessary_columns['physical_users'][0]]),
                             pd.Series(data[necessary_columns['physical_users'][1]]),
                             pd.Series(data[necessary_columns['physical_users'][2]]),
                             pd.Series(data[necessary_columns['physical_users'][3]])
                             ])
        data = data.transpose()
    elif len(set(necessary_columns['legal_users']) & set(data)) == 4:
        data = data[(data['Группа потребителей'] == users[0]) |
                    (data['Группа потребителей'] == users[1]) |
                    (data['Группа потребителей'] == users[2]) |
                    (data['Группа потребителей'] == users[3]) |
                    (data['Группа потребителей'] == users[4]) |
                    (data['Группа потребителей'] == users[5])
                    ]
        data = pd.DataFrame([pd.Series(data[necessary_columns['legal_users'][0]]),
                             pd.Series(data[necessary_columns['legal_users'][1]]),
                             pd.Series(data[necessary_columns['legal_users'][2]]),
                             pd.Series(data[necessary_columns['legal_users'][3]])
                             ])
        data = data.transpose()
    else:
        return "Error. Wrong list"
    return data


def brute_force(index, columns):
    """Функция для поочередного перебора значений в массиве данных"""
    if columns > 50:
        columns = 50
    for i in range(index):
        for j in range(columns):
            yield i, j


def enter_information(base, data, unreported):
    """Функция заполняет базу отсортированными данными"""
    position_district = find_position_district(base, data)
    for i in range(len(data.index)):
        if type(data.iloc[i, 1]) != str:
            continue
        address_user = change_line(data.iloc[i, 1], True)
        flag = False
        for j in position_district:
            flag1 = False
            if type(base.iloc[j, 2]) == list and len(base.iloc[j, 2]) > 0:
                for t in base.iloc[j, 2]:
                    if t in address_user:
                        flag1 = True
                        break
                if flag1:
                    continue
            for g in base.iloc[j, 7]:
                if g in address_user:
                    try:
                        base.iloc[j, 4] += int(data.iloc[i, 3])
                    except Exception:
                        pass
# TODO Выводить на экран ячейку которую не смог посчитать в сумму
                    base.iloc[j, 3].add(data.iloc[i, 2])
                    data.iloc[i, 0] = None
                    flag = True
                    break
            if flag:
                break
        if data.iloc[i, 0] is not None:
            unreported.append({'РЭС': data.iloc[i, 0], 'Адреc': data.iloc[i, 1],
                               'Номер счетчика': data.iloc[i, 2], 'Объем': data.iloc[i, 3]})
    return base, unreported


def find_position_district(base, data):
    district = change_line(data.iloc[0, 0])
    district = district.split()[0]
    position = []
    if district == 'сыктывдинский' or district == 'сыктывкарский' or district == 'эжвинский':
        position = [i for i in range(189, 236)] + [i for i in range(447, 454)]
    else:
        for i in range(len(base.index)):
            if base.iloc[i, 0] is None:
                continue
            elif district in base.iloc[i, 0]:
                position.append(i)
    return position


def make_excel():
    """Создает  Excel файл куда вносится информация"""
    wb = openpyxl.Workbook()
    ws_in = wb.active
    wb.create_sheet('Неучтенные адреса')
    ws_in.title = 'Главная книга'
    name_book = 'Отчет'

    base = openpyxl.load_workbook(os.path.abspath('baza.xlsx'))
    ws_out = base.worksheets[0]
    for row in range(1, ws_out.max_row + 1):
        ws_in['A{0}'.format(row)] = ws_out.cell(row, 1).value
        ws_in['B{0}'.format(row)] = ws_out.cell(row, 2).value
        ws_in['C{0}'.format(row)] = ws_out.cell(row, 3).value
    ws_in['D1'] = 'Объем, кВт'
    ws_in['E1'] = 'Кол-во точек'
    ws_in.column_dimensions['A'].width = 40
    ws_in.column_dimensions['B'].width = 16
    ws_in.column_dimensions['C'].width = 23
    ws_in.column_dimensions['D'].width = 10
    ws_in.column_dimensions['E'].width = 15
    ws_in_unreported = wb['Неучтенные адреса']
    ws_in_unreported['A1'] = 'РЭС'
    ws_in_unreported['B1'] = 'Адрес'
    ws_in_unreported['C1'] = 'Номер счетчика'
    ws_in_unreported['D1'] = 'Объем'
    ws_in_unreported.column_dimensions['A'].width = 15
    ws_in_unreported.column_dimensions['B'].width = 80
    ws_in_unreported.column_dimensions['C'].width = 20
    ws_in_unreported.column_dimensions['D'].width = 10
    wb.save('{}.xlsx'.format(name_book))
    return name_book


def enter_in_excel(wb_in, data_out, data_unreported):
    """Вносит всю подготовленную информацию в созданный Excel"""
    wb = openpyxl.load_workbook(os.path.abspath('{}.xlsx'.format(wb_in)))
    ws_in = wb.worksheets[0]
    for row in range(2, ws_in.max_row + 1):
        if data_out.iloc[row - 2, 0] is None:
            continue
        ws_in['D{0}'.format(row)] = data_out.iloc[row - 2, 4]
        data_out.iloc[row - 2, 3] = len(data_out.iloc[row - 2, 3])
        ws_in['E{0}'.format(row)] = data_out.iloc[row - 2, 3]
    ws_in_unreported = wb.worksheets[1]
    for row in range(0, len(data_unreported.index)):
        ws_in_unreported['A{0}'.format(row + 2)] = data_unreported.iloc[row, 0]
        ws_in_unreported['B{0}'.format(row + 2)] = data_unreported.iloc[row, 1]
        ws_in_unreported['C{0}'.format(row + 2)] = data_unreported.iloc[row, 2]
        ws_in_unreported['D{0}'.format(row + 2)] = data_unreported.iloc[row, 3]
    wb.save('{}.xlsx'.format(wb_in))
    return


def reanalysis(base, unreported):
    """Повторно анализирует неучтенные адреса. Уже простым перебором без учета районного центра"""
    new_unreported = []
    for i in range(len(unreported.index)):
        address_user = change_line(unreported.iloc[i, 1], True)
        for j in range(len(base.index)):
            if base.iloc[j, 0] is None:
                continue
            flag = False
            for town in base.iloc[j, 7]:
                if town in address_user:
                    base.iloc[j, 4] += int(unreported.iloc[i, 3])
                    base.iloc[j, 3].add(unreported.iloc[i, 2])
                    unreported.iloc[i, 0] = None
                    flag = True
                    break
            if flag:
                break
        if unreported.iloc[i, 0] is not None:
            new_unreported.append({'РЭС': unreported.iloc[i, 0], 'Адреc': unreported.iloc[i, 1],
                                   'Номер счетчика': unreported.iloc[i, 2], 'Объем': unreported.iloc[i, 3]})
    return base, new_unreported


print(colorama.Fore.RED + 'Запуск программы')
print('Программа запущена. Version 1.17')
print(colorama.Style.RESET_ALL)
key_input = input('''Введите:
      help для для получения справки о программе
      n для выхода из программы
      любую другую кнопку (или просто нажмите Enter для запуска программы\n''')
if key_input == 'help':
    print('''Программа будет запрашивать путь к данным. необходимо указывать путь к папке 
    где хранятся все данные которые хотите посчитать. 
    !!Если в указанной папке файлы с данными будут находиться 
    еще во внутренних папках, то программа их не увидит!!''', end='\n\n')
    print('''Программа пока не может:'
           1. Есть н.п у которых одинаковое название (Например п. Ягдор), программа пока не может их разделять и скорее
           всего занесет сразу все значения только одному н.п. Как и в старой программе такие города нужно 
           откорректировать в ручную. Правильное функционирование (если получиться) будет в марте 2021 года.
           
           в Version 1.17 добавлена функция сбора неутченных адресов.
           "раздел будет дополняться по мере необходимости"
           
           ''')
    input('Нажмите любую кнопку для начала работы программы')
elif key_input == 'n':
    sys.exit()

print('''Обращаю ваше внимание, что из-за неправильного создания Excel файла с огромным количеством лишних пустых 
строк и столбцов будет значительно увеличина длительность выполнения программы.
К примеру такие Excel вместо пары Мб могут весить десятки и сотни... 
Но надеюсь в этот раз обойдется без таких проблемных файлов.
Теперь можно налить себе кружку чая, программа готова и начинает свою работу''')

# Открываем и подготоавливаем файл куда будет вносится информация
base_for_fill = main_file()
print('Файл для записи подготовлен')

# Находим нужные файлы с необходимой информацией
files_for_import = find_files_with_information()
print('Информация найдена')
print('Запушен процесс подготовки инфомации')
start = time.time()
unreported_address = []
# Перебор файлов и импорт нужной информации

for file in files_for_import:
    base_for_fill, unreported_address = import_information(base_for_fill, file, unreported_address)
    gc.collect()
unreported_address = pd.DataFrame(unreported_address)
print('Вся информация подготовлена')
print(colorama.Fore.GREEN +
      'Обращаем внимание, что программой необработанно {0} адресов'.format(len(unreported_address.index)))
print('В основном это связано с некорректным заполнением адреса ')
print('Но бывают случаи когда программа ошибочно не обрабатывает адрес')
print('Это обычно связано когда название н.п. входит в название районного центра (Пример "Поруб" и "СпасПоруб")')
print(colorama.Style.BRIGHT +
      'Для решение этой проблему предлагаю повторно проанализировать только неучтенные адреса по другому алгоритму')
key_input = input('Для повторного анализа введите да \nДля пропуска данного пункта и продолжении программы')
print(colorama.Style.RESET_ALL)
if key_input == 'ДА' or key_input == 'Да' or key_input == 'да' or key_input == 'дА':
    print('Начат повторный анализ')
    base_for_fill, unreported_address = reanalysis(base_for_fill, unreported_address)
    unreported_address = pd.DataFrame(unreported_address)
    print('Повторный анализ завершен. Осталось {0} неучтенных адресов'.format(len(unreported_address.index)))
else:
    pass
print('Все неучтенные адреса храняться в файле Отчет.xlsx, Лист "Неучтенные адреса"')
print('Начата запись данных в итоговый файл')
sheet_in = make_excel()
enter_in_excel(sheet_in, base_for_fill, unreported_address)
minute = int((time.time() - start) // 60)
second = int((time.time() - start) % 60)
print('Программа завершила расчет. Время выполнения {0} мин. {1} c.'.format(minute, second))
print(colorama.Style.BRIGHT + 'Вся информация внесена в файл Отчет.xlsx')
print('Данный файл находится в каталоге программы')
input('Введите любую кнопку для завершения программы')
